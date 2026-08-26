"""
Catatan Kas - Rekap Pembayaran Perorang
Sumber logika rekap: periode pembayaran (Dari Bulan -> Sampai Bulan),
bukan tanggal transaksi. Dipakai untuk ekspor/renderer Android.
"""
from __future__ import annotations

from collections import defaultdict
from typing import Any, Iterable
import base64
import io
import json

BULAN = [
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember"
]

_ALIASES_BULAN = {
    "jan": 1, "januari": 1, "feb": 2, "februari": 2,
    "mar": 3, "maret": 3, "apr": 4, "april": 4,
    "mei": 5, "may": 5, "jun": 6, "juni": 6,
    "jul": 7, "juli": 7, "agu": 8, "ags": 8, "agustus": 8,
    "aug": 8, "sep": 9, "september": 9, "okt": 10, "oktober": 10,
    "nov": 11, "november": 11, "des": 12, "desember": 12,
}


def _norm(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _bulan(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        n = int(value)
        return n if 1 <= n <= 12 else (n + 1 if 0 <= n <= 11 else None)
    s = _norm(value)
    if s.isdigit():
        n = int(s)
        return n if 1 <= n <= 12 else (n + 1 if 0 <= n <= 11 else None)
    return _ALIASES_BULAN.get(s)


def _ambil_nama(t: dict[str, Any]) -> str:
    for key in (
        "nama", "namaSantri", "santriNama", "namaSiswaSiswi",
        "namaSiswa-Siswi", "nama_siswa_siswi", "name", "santri",
        "keterangan",
    ):
        value = t.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return "Tanpa Nama"


def _jenis(t: dict[str, Any]) -> str:
    value = t.get("jenis", t.get("kategori", t.get("jenisPemasukan", t.get("jenisPembayaran", t.get("tipe", "Lainnya")))) )
    return _norm(value) or "lainnya"


def _jumlah(t: dict[str, Any]) -> float:
    value = t.get("nominal", t.get("jumlah", t.get("nilai", t.get("total", t.get("amount", t.get("nominalPemasukan", t.get("jumlahBayar", 0)))))))
    if isinstance(value, str):
        value = value.replace("Rp", "").replace(" ", "").replace(".", "").replace(",", ".")
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _tahun(t: dict[str, Any]) -> int | None:
    for key in ("tahun", "tahunPembayaran", "tahunPeriode", "tahunKewajiban", "tahunBayar"):
        try:
            if t.get(key) not in (None, ""):
                return int(t[key])
        except (TypeError, ValueError):
            pass
    return None


def _periode_transaksi(t: dict[str, Any]) -> list[int]:
    a = None
    b = None
    for key in ("bulanDari", "bulanMulai", "bulanAwal", "dariBulan", "periodeDari", "periodeWajibDari"):
        a = _bulan(t.get(key))
        if a is not None:
            break
    for key in ("bulanSampai", "bulanAkhir", "bulanAkhirPembayaran", "sampaiBulan", "periodeSampai", "periodeAkhir", "periodeWajibSampai"):
        b = _bulan(t.get(key))
        if b is not None:
            break
    if a is None and b is None:
        return []
    if a is None:
        a = b
    if b is None:
        b = a
    if a > b:
        a, b = b, a
    return list(range(a, b + 1))


def _jenis_sama(a: str, b: str) -> bool:
    alias = {"syahriah": "syahriyyah", "syahriyyah": "syahriyyah"}
    return alias.get(_norm(a), _norm(a)) == alias.get(_norm(b), _norm(b))


def rekap_pembayaran(
    transaksi: Iterable[dict[str, Any]],
    tahun: int | None = None,
    jenis: str = "Semua",
    peserta: Iterable[dict[str, Any]] | None = None,
    periode_wajib: dict[str, dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """Rekap berdasarkan periode pembayaran, bukan tanggal transaksi."""
    paid: dict[str, set[int]] = defaultdict(set)
    total: dict[str, float] = defaultdict(float)
    names: set[str] = set()

    if peserta:
        for p in peserta:
            n = _ambil_nama(p)
            if n != "Tanpa Nama":
                names.add(n)

    for t in transaksi:
        if not isinstance(t, dict):
            continue
        n = _ambil_nama(t)
        if n == "Tanpa Nama":
            continue
        ty = _jenis(t)
        if jenis and jenis != "Semua" and not _jenis_sama(ty, jenis):
            continue
        ty_year = _tahun(t)
        if tahun is not None and ty_year not in (None, int(tahun)):
            continue
        names.add(n)
        bulan = _periode_transaksi(t)
        paid[n].update(bulan)
        total[n] += _jumlah(t)

    hasil = []
    for nama in sorted(names, key=str.casefold):
        p = (periode_wajib or {}).get(_norm(nama), {})
        start = _bulan(p.get("periodeWajibDari", p.get("dariBulan", p.get("bulanMulai"))))
        end = _bulan(p.get("periodeWajibSampai", p.get("sampaiBulan", p.get("bulanSelesai"))))
        if start is not None and end is not None and start > end:
            start, end = end, start

        row = {"Santri": nama}
        wajib = set(range(start, end + 1)) if start is not None and end is not None else set()
        for nomor, bulan in enumerate(BULAN, 1):
            if nomor not in wajib:
                row[bulan] = "—"
            elif nomor in paid[nama]:
                row[bulan] = "✓"
            else:
                row[bulan] = "✗"

        sudah = sorted(wajib.intersection(paid[nama]))
        row["Periode Wajib"] = f"{BULAN[start - 1]} – {BULAN[end - 1]}" if wajib else "Belum ditentukan"
        row["Jumlah Bulan"] = len(wajib)
        row["Sudah Bayar"] = len(sudah)
        row["Sudah Bayar Sampai"] = BULAN[sudah[-1] - 1] if sudah else "Belum bayar"
        row["Total Pembayaran"] = total[nama]
        hasil.append(row)
    return hasil


def simpan_excel(rekap: list[dict[str, Any]], path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Pembayaran"
    headers = ["Santri", *BULAN, "Periode Wajib", "Jumlah Bulan", "Sudah Bayar Sampai", "Total Pembayaran"]
    ws.append(headers)
    for row in rekap:
        ws.append([row.get(h, "") for h in headers])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        letter = col[0].column_letter
        ws.column_dimensions[letter].width = min(max(max(len(str(c.value or "")) for c in col) + 2, 10), 24)
    wb.save(path)


def export_excel_base64(payload_json: str) -> str:
    payload = json.loads(payload_json or "{}")
    transaksi = payload.get("transaksi", [])
    tahun = payload.get("tahun")
    tahun = int(tahun) if tahun not in (None, "") else None
    hasil = rekap_pembayaran(
        transaksi,
        tahun,
        payload.get("jenis", "Semua"),
        payload.get("peserta", []),
        payload.get("periodeWajib", {}),
    )
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Pembayaran"
    headers = ["Santri", *BULAN, "Periode Wajib", "Jumlah Bulan", "Sudah Bayar Sampai", "Total Pembayaran"]
    ws.append(headers)
    for row in hasil:
        ws.append([row.get(h, "") for h in headers])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        letter = col[0].column_letter
        ws.column_dimensions[letter].width = min(max(max(len(str(c.value or "")) for c in col) + 2, 10), 24)
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode("ascii")
